import os
import time
import requests
import openpyxl
import pandas as pd

current_path = os.getcwd()
print("Your current Path right now is : ", current_path)

xlsx_file = input("Please insert path to your excel (.xlsx) files here..\nExample like : C:/Users/MuhamadHafiz_gbgcadq/Desktop/msisdn.xlsx\n\nPath : \n")

workbook = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
worksheet = workbook.active


celcom_url = "https://igw.apistg.celcom.com.my/oneapi/queryprofile/v1/"
celcom_url_type1 = "/subscribertype"
#smart_url = ""
token = input("Please insert working access token here : \n")


df = []
for i in range(1, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        msisdn = col[i].value
        URL = str(celcom_url) + str(msisdn) + str(celcom_url_type1)
        response = requests.get(URL, headers={"Accept": "application/x-www-form-urlencoded", "Accept-Charset": "UTF-8", "Authorization": "Bearer " + token + '"', "Cache-Control": "no-cache"})
        res = response.text

##Transformation

        data = pd.DataFrame({'MSISDN':[msisdn], 'Result':[res]}).dropna()
        df.append(data)
        time.sleep(1) ##Sleep for 1 second to avoid TPS Throttling issues
data = pd.concat(df, ignore_index=True)
df1 = data
print(df1)
df1.to_excel(str(current_path) + "\result.xlsx", encoding='utf-8', index=False)
exit()